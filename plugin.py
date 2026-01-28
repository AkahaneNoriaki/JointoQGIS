# -*- coding: utf-8 -*-
"""
ExcelCSVLatestJoinAllInOne
- Select Excel(xlsx) + sheet
- Export to CSV (latest.csv + fallback.csv) into chosen folder
- Create/refresh a virtual join to a target layer (remove old joins first)
- Auto sync by polling Excel mtime/size
Tested target: QGIS 3.34 + Python 3.12 on Windows
"""
from __future__ import annotations

import os
import re
import csv
import time
import traceback
import hashlib
from typing import List, Optional

from qgis.PyQt.QtCore import Qt, QTimer, QSettings
from qgis.PyQt.QtWidgets import (
    QDockWidget, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QComboBox, QCheckBox, QSpinBox,
    QFileDialog, QMessageBox
)
from qgis.core import (
    QgsProject, QgsVectorLayer, QgsVectorLayerJoinInfo,
    QgsVectorFileWriter, QgsCoordinateTransformContext,
    QgsProviderRegistry, QgsEditorWidgetSetup
)

try:
    import openpyxl  # type: ignore
    _HAVE_OPENPYXL = True
except Exception:
    openpyxl = None
    _HAVE_OPENPYXL = False


class ExcelCSVLatestJoinAllInOne:
    PLUGIN_NAME = "JointoQGIS"
    ORG = "excelcsv_latest_join_allinone"

    PROJECT_NS = "excelcsv_latest_join_allinone/project"
    SETTINGS_NS = "excelcsv_latest_join_allinone/ui"

    def __init__(self, iface):
        self.iface = iface
        self.action = None
        self.dock: Optional[QDockWidget] = None

        # global UI settings (persist across projects)
        self._qsettings = QSettings()
        self.persist_enabled = False  # per-project only (default unchecked)
        self.auto_show_dock: bool = bool(self._qsettings.value(f"{self.SETTINGS_NS}/auto_show_dock", False, type=bool))

        # paths / settings
        self.excel_path: str = ""
        self.sheet_name: str = ""
        self.csv_folder: str = ""

        self.latest_csv: str = ""
        self.fallback_csv: str = ""

        # keep reference to loaded CSV layer (avoid GC / ensure join fields visible)
        self._csv_layer = None

        # sync state
        self._timer = QTimer()
        self._timer.setInterval(2000)
        self._timer.timeout.connect(self._poll_updates)


        self._pending_export = False
        self._export_retry_count = 0
        self._export_retry_max = 12  # up to ~30s backoff
        self._last_excel_mtime: float = 0.0
        self._last_excel_size: int = -1

        self._last_csv_mtime: float = 0.0
        self._last_csv_size: int = -1

        # join config
        self.target_layer_id: str = ""
        self.layer_join_field: str = ""
        self.csv_join_field: str = ""

        # ui refs
        self._lbl_status = None
        self._cmb_sheet = None
        self._cmb_layer = None
        self._cmb_layer_key = None
        self._cmb_csv_key = None
        self._spin_interval = None
        self._chk_auto = None
        self._chk_auto_show = None

    # ---------------- QGIS plugin boilerplate ----------------
    def _restore_after_ui_ready(self) -> None:
        """Restore saved settings after UI combos are populated (retry)."""
        # prevent autosave recursion while restoring UI
        self._is_restoring = True
        try:
            try:
                self._restore_tries = getattr(self, "_restore_tries", 0) + 1
            except Exception:
                self._restore_tries = 1

            # Push saved paths into UI
            try:
                self._txt_excel.setText(self.excel_path or "")
            except Exception:
                pass
            try:
                self._txt_csv.setText(self.csv_folder or "")
            except Exception:
                pass
            try:
                self._apply_csv_folder_runtime()
            except Exception:
                pass
            try:
                self._spin_interval.setValue(30 if self._get_interval_sec() < 45 else 60)
            except Exception:
                pass
            try:
                self._chk_auto_show.setChecked(bool(getattr(self, "auto_show_dock", False)))
            except Exception:
                pass

            # Populate combos (needs UI built)
            try:
                self._populate_sheets()
            except Exception:
                pass
            try:
                self._populate_layers()
            except Exception:
                pass
            try:
                self._restore_layer_selection_c_deferred(tries=8)
            except Exception:
                pass

            # Apply selections
            try:
                if getattr(self, "sheet_name", "") and self._cmb_sheet.count() > 0:
                    if self._cmb_sheet.findText(self.sheet_name) >= 0:
                        self._cmb_sheet.setCurrentText(self.sheet_name)
            except Exception:
                pass
            try:
                if getattr(self, "layer_name", "") and self._cmb_layer.count() > 0:
                    if self._cmb_layer.findText(self.layer_name) >= 0:
                        self._cmb_layer.setCurrentText(self.layer_name)
            except Exception:
                pass
            try:
                if getattr(self, "layer_join_field", "") and self._cmb_layer_key.count() > 0:
                    if self._cmb_layer_key.findText(self.layer_join_field) >= 0:
                        self._cmb_layer_key.setCurrentText(self.layer_join_field)
            except Exception:
                pass
            try:
                if getattr(self, "csv_join_field", "") and self._cmb_csv_key.count() > 0:
                    if self._cmb_csv_key.findText(self.csv_join_field) >= 0:
                        self._cmb_csv_key.setCurrentText(self.csv_join_field)
            except Exception:
                pass

            # Retry if still empty
            try:
                need_retry = (self._cmb_layer.count() == 0) or (self._cmb_sheet.count() == 0)
            except Exception:
                need_retry = False

            if need_retry and self._restore_tries < 8:
                try:
                    QTimer.singleShot(250, self._restore_after_ui_ready)
                except Exception:
                    pass
            else:
                try:
                    self._log("project: settings restored")
                except Exception:
                    pass
        finally:
            try:
                self._is_restoring = False
            except Exception:
                pass
    def _connect_project_signals(self) -> None:
        if getattr(self, '_project_signals_connected', False):
            return
        self._project_signals_connected = True
        prj = QgsProject.instance()
        # Read/open project
        try:
            if hasattr(prj, 'projectRead'):
                prj.projectRead.connect(self._on_project_read)
            elif hasattr(prj, 'readProject'):
                prj.readProject.connect(self._on_project_read)
        except Exception:
            pass
        # Saved
        try:
            # Prefer writeProject (fires BEFORE writing the project file) so entries persist in .qgs/.qgz
            if hasattr(prj, 'writeProject'):
                prj.writeProject.connect(self._on_project_saved)
            elif hasattr(prj, 'projectSaved'):
                prj.projectSaved.connect(self._on_project_saved)
        except Exception:
            pass
        # Cleared/new project (supports switching without restart)
        try:
            if hasattr(prj, 'cleared'):
                prj.cleared.connect(self._on_project_cleared)
        except Exception:
            pass

    def _on_project_read(self, *args) -> None:
        """Project opened: restore settings if project has persist_enabled=1."""
        try:
            use_project = False
            try:
                use_project = bool(self._project_has_persist_flag())
            except Exception:
                use_project = False
            self.persist_enabled = bool(use_project)
            # update checkbox without emitting signals
            try:
                if hasattr(self, '_chk_persist'):
                    self._set_chk_checked_blocked(self._chk_persist, self.persist_enabled)
            except Exception:
                pass
            if use_project:
                try:
                    self._load_project_settings()
                    self._apply_csv_folder_runtime()
                    self._log('project: settings restored')
                except Exception as e:
                    try:
                        self._log(f'project: restore failed: {e}')
                    except Exception:
                        pass
                try:
                    if getattr(self, 'dock', None) is not None:
                        QTimer.singleShot(0, self._restore_after_ui_ready)
                except Exception:
                    pass
        except Exception:
            pass

    def _on_project_cleared(self, *args) -> None:
        """New/empty project created (or cleared): reset UI state without dirtying."""
        try:
            self.persist_enabled = False
            try:
                if hasattr(self, '_chk_persist'):
                    self._set_chk_checked_blocked(self._chk_persist, False)
            except Exception:
                pass
            # Clear UI only; do not touch project properties or globals
            try:
                if getattr(self, 'dock', None) is not None:
                    self._is_restoring = True
                    try:
                        self._txt_excel.setText('')
                        self._txt_csv.setText('')
                    except Exception:
                        pass
                    self._is_restoring = False
            except Exception:
                pass
        except Exception:
            pass


    def _on_project_saved(self, *args) -> None:
        if bool(getattr(self, '_is_restoring', False)):
            return
        save_phase = 'writeProject' if len(args) else 'projectSaved'
        try:
            self._save_project_settings()
            self._log(f'project: settings saved ({save_phase})')
        except Exception as e:
            try:
                self._log(f'project: save failed: {e}')
            except Exception:
                pass

    def _apply_restored_selections(self) -> None:
        try:
            if getattr(self, "layer_name", ""):
                idx = self.cbo_layer.findText(self.layer_name)
                if idx >= 0:
                    self.cbo_layer.setCurrentIndex(idx)
            if getattr(self, "layer_join_field", ""):
                idx = self.cbo_layer_field.findText(self.layer_join_field)
                if idx >= 0:
                    self.cbo_layer_field.setCurrentIndex(idx)
            if getattr(self, "csv_join_field", ""):
                idx = self.cbo_csv_field.findText(self.csv_join_field)
                if idx >= 0:
                    self.cbo_csv_field.setCurrentIndex(idx)
        except Exception:
            pass

    def _wire_autosave(self) -> None:
        if getattr(self, "_autosave_wired", False):
            return
        self._autosave_wired = True

        def _save_now(*args):
            if not self._project_can_persist():
                return

            if bool(getattr(self, '_is_restoring', False)):
                return

            try:
                self._save_project_settings()
            except Exception:
                pass

        for w, sig in [
            ("txt_excel", "textChanged"),
            ("txt_csv_folder", "textChanged"),
            ("cbo_sheet", "currentTextChanged"),
            ("cbo_layer", "currentTextChanged"),
            ("cbo_layer_field", "currentTextChanged"),
            ("cbo_csv_field", "currentTextChanged"),
        ]:
            try:
                getattr(getattr(self, w), sig).connect(_save_now)
            except Exception:
                pass
        try:
            self.spin_interval.valueChanged.connect(_save_now)
        except Exception:
            pass
        try:
            self.chk_autoshow.stateChanged.connect(_save_now)
        except Exception:
            pass

    def initGui(self):
        from qgis.PyQt.QtWidgets import QAction
        self.action = QAction("Excel/CSV 同期", self.iface.mainWindow())
        self.action.triggered.connect(self.show_dock)
        self.iface.addPluginToMenu(self.PLUGIN_NAME, self.action)
        self.iface.addToolBarIcon(self.action)

        # Connect project signals EARLY (important when QGIS opens a project before plugins finish loading)
        try:
            self._connect_project_signals()
        except Exception:
            pass
        try:
            self._wire_autosave()
        except Exception:
            pass

        # If a project is already open, sync persist state now
        try:
            self._on_project_read()
        except Exception:
            pass

        # auto show dock on QGIS start (user setting)
        if self.auto_show_dock:
            try:
                self.show_dock()
            except Exception:
                pass

        # --- aliases for persistence helpers ---
        try:
            self.txt_excel = self._txt_excel
            self.txt_csv_folder = self._txt_csv
            self.cbo_sheet = self._cmb_sheet
            self.cbo_layer = self._cmb_layer
            self.cbo_layer_field = self._cmb_layer_key
            self.cbo_csv_field = self._cmb_csv_key
            self.spin_interval = self._spin_interval
            self.chk_autoshow = self._chk_auto_show
        except Exception:
            pass

    def unload(self):

        try:
            if self.action:
                self.iface.removePluginMenu(self.PLUGIN_NAME, self.action)
                self.iface.removeToolBarIcon(self.action)
        except Exception:
            pass
        self._timer.stop()
        if self.dock:
            self.iface.mainWindow().removeDockWidget(self.dock)
            self.dock = None

    # ---------------- UI ----------------
    def show_dock(self):
        if self.dock is None:
            self.dock = QDockWidget(self.PLUGIN_NAME, self.iface.mainWindow())
            self.dock.setObjectName("ExcelCSVLatestJoinAllInOneDock")
            w = QWidget()
            root = QVBoxLayout(w)

            # 0) Excel -> CSV
            root.addWidget(QLabel("0) Excel → CSV (latest.csv / fallback.csv)"))

            row = QHBoxLayout()
            btn_excel = QPushButton("Excel(xlsx)を選択")
            btn_excel.clicked.connect(self.select_excel)
            self._txt_excel = QLineEdit()
            self._txt_excel.setReadOnly(True)
            row.addWidget(btn_excel)
            row.addWidget(self._txt_excel, 1)
            root.addLayout(row)

            row = QHBoxLayout()
            row.addWidget(QLabel("シート"))
            self._cmb_sheet = QComboBox()
            self._cmb_sheet.currentTextChanged.connect(self._on_sheet_changed)
            row.addWidget(self._cmb_sheet, 1)
            root.addLayout(row)

            row = QHBoxLayout()
            btn_csv = QPushButton("CSVフォルダを選択")
            btn_csv.clicked.connect(self.select_csv_folder)
            self._txt_csv = QLineEdit()
            self._txt_csv.setReadOnly(True)
            row.addWidget(btn_csv)
            row.addWidget(self._txt_csv, 1)
            root.addLayout(row)

            row = QHBoxLayout()
            btn_export = QPushButton("Excel→CSV生成（手動）")
            btn_export.clicked.connect(lambda: self._schedule_export("manual"))
            row.addWidget(btn_export)
            btn_open = QPushButton("CSVを開く")
            btn_open.clicked.connect(self._open_latest_csv)
            row.addWidget(btn_open)
            root.addLayout(row)

            self._lbl_status = QLabel("状態: -")
            root.addWidget(self._lbl_status)

            # 1) Join config
            root.addSpacing(8)
            root.addWidget(QLabel("1) 自動結合設定（仮Join）"))

            row = QHBoxLayout()
            row.addWidget(QLabel("対象レイヤ"))
            self._cmb_layer = QComboBox()
            self._cmb_layer.currentIndexChanged.connect(self._on_layer_changed)
            btn_refresh = QPushButton("レイヤ一覧更新")
            btn_refresh.clicked.connect(self._populate_layers)
            row.addWidget(self._cmb_layer, 1)
            row.addWidget(btn_refresh)
            root.addLayout(row)

            row = QHBoxLayout()
            row.addWidget(QLabel("レイヤ Joinキー"))
            self._cmb_layer_key = QComboBox()
            self._cmb_layer_key.currentTextChanged.connect(self._on_join_keys_changed)
            row.addWidget(self._cmb_layer_key, 1)
            row.addWidget(QLabel("CSV Joinキー"))
            self._cmb_csv_key = QComboBox()
            self._cmb_csv_key.currentTextChanged.connect(self._on_join_keys_changed)
            row.addWidget(self._cmb_csv_key, 1)
            root.addLayout(row)

            row = QHBoxLayout()
            btn_join_now = QPushButton("今すぐJoin更新")
            btn_join_now.clicked.connect(lambda: self.execute_virtual_join(quiet=False))
            row.addWidget(btn_join_now)
            btn_clear_join = QPushButton("Join解除")
            btn_clear_join.clicked.connect(self.remove_existing_joins)
            row.addWidget(btn_clear_join)
            root.addLayout(row)

            # 2) Auto sync
            root.addSpacing(8)
            root.addWidget(QLabel("2) 自動同期(Excel更新→CSV生成→CSV検知→自動結合)"))

            row = QHBoxLayout()
            self._chk_auto_show = QCheckBox("QGIS起動時にこのドックを自動表示")
            self._chk_auto_show.setChecked(self.auto_show_dock)
            self._chk_auto_show.stateChanged.connect(self._on_auto_show_changed)
            root.addWidget(self._chk_auto_show)

            self._chk_persist = QCheckBox("設定を保存して次回復元")
            # init persist checkbox from project flag (do not emit stateChanged)
            try:
                self.persist_enabled = bool(self._project_has_persist_flag())
            except Exception:
                self.persist_enabled = False
            self._set_chk_checked_blocked(self._chk_persist, self.persist_enabled)
            def _on_persist_changed(state: int):
                self.persist_enabled = (state == Qt.Checked)
                # Project-only persistence
                try:
                    if self.persist_enabled:
                        self._save_project_settings()
                        try:
                            self._qs_prj_set('persist_enabled', '1')
                        except Exception:
                            pass
                        self._log('project: persist enabled')
                    else:
                        self._prj_set('persist_enabled', '0')
                        try:
                            self._qs_prj_set('persist_enabled', '0')
                        except Exception:
                            pass
                        self._log('project: persist disabled')
                except Exception:
                    pass
            self._chk_persist.stateChanged.connect(_on_persist_changed)
            root.addWidget(self._chk_persist)

            self._chk_auto = QCheckBox("自動同期ON")
            self._chk_auto.setChecked(True)
            self._chk_auto.stateChanged.connect(self._on_auto_changed)
            row.addWidget(self._chk_auto)

            row.addWidget(QLabel("監視間隔"))
            self._spin_interval = QSpinBox()
            self._spin_interval.setRange(30, 60)
            self._spin_interval.setValue(30)
            self._spin_interval.valueChanged.connect(self._on_interval_changed)
            row.addWidget(self._spin_interval)
            row.addWidget(QLabel("秒"))
            root.addLayout(row)

            self.dock.setWidget(w)
            self.iface.mainWindow().addDockWidget(Qt.RightDockWidgetArea, self.dock)
            # load settings (project only when opted-in)
            try:
                v = self._prj_get('persist_enabled', '0')
            except Exception:
                v = None
            use_project = str(v) in ('1', 'true', 'True')
            if use_project:
                self.persist_enabled = True
                try:
                    if hasattr(self, '_chk_persist'):
                        self._set_chk_checked_blocked(self._chk_persist, True)
                except Exception:
                    pass
                try:
                    self._load_project_settings()
                    self._log('project: settings restored')
                except Exception:
                    pass
            else:
                self.persist_enabled = False
                try:
                    if hasattr(self, '_chk_persist'):
                        self._set_chk_checked_blocked(self._chk_persist, False)
                except Exception:
                    pass

            # init lists
            self._populate_layers()
            try:
                self._apply_state_to_ui()
            except Exception:
                pass
            self._update_status("起動")

        self.dock.show()
        self.dock.raise_()

        # Restore settings from the currently opened project (supports switching projects without restarting QGIS)
        try:
            if self._project_has_persist_flag():
                self.persist_enabled = True
                if hasattr(self, '_chk_persist'):
                    self._set_chk_checked_blocked(self._chk_persist, True)
                try:
                    self._load_project_settings()
                    self._log('project: settings restored')
                except Exception:
                    pass
            else:
                self.persist_enabled = False
                if hasattr(self, '_chk_persist'):
                    self._set_chk_checked_blocked(self._chk_persist, False)
        except Exception:
            pass

        # Apply in-memory state to UI
        try:
            QTimer.singleShot(0, self._restore_after_ui_ready)
        except Exception:
            try:
                self._restore_after_ui_ready()
            except Exception:
                pass

        # start timer if auto
        if self._chk_auto is not None and self._chk_auto.isChecked():
            if not self._timer.isActive():
                self._timer.start()

    def _on_auto_show_changed(self, state: int):
        self.auto_show_dock = (state == Qt.Checked)
        try:
            self._qsettings.setValue(f"{self.SETTINGS_NS}/auto_show_dock", self.auto_show_dock)
        except Exception:
            pass
    # ---------------- Global persistence (QSettings) ----------------
    def _save_global_settings(self) -> None:
        # disabled: project-only persistence
        return

    def _load_global_settings(self) -> None:
        # disabled: project-only persistence
        return

    def _on_project_read(self):
        self._load_project_settings()
        if self.dock is not None:
            try:
                self._apply_state_to_ui()
            except Exception:
                pass
        if self.auto_show_dock:
            try:
                self.show_dock()
            except Exception:
                pass

    def _on_project_write(self):
        try:
            self._save_project_settings()
        except Exception:
            pass
        try:
            self._save_global_settings()
        except Exception:
            pass

    def _project_can_persist(self) -> bool:
        """Only write custom properties when project is a saved file and persistence is enabled."""
        try:
            prj = QgsProject.instance()
            has_file = bool(prj.fileName() or "")
        except Exception:
            has_file = False
        try:
            enabled = bool(getattr(self, "persist_enabled", False))
        except Exception:
            enabled = False
        return bool(has_file and enabled)

    def _project_has_persist_flag(self) -> bool:
        """True if current project has persist_enabled=1."""
        try:
            v = self._prj_get('persist_enabled', '0')
        except Exception:
            v = None
        ok = str(v) in ('1','true','True')
        if ok:
            return True
        try:
            vv = self._qs_prj_get('persist_enabled', '0')
            return str(vv) in ('1','true','True')
        except Exception:
            return False

    def _proj_id(self) -> str:
        """Stable id from current project file path; empty if project not saved."""
        try:
            p = (QgsProject.instance().fileName() or '').strip()
        except Exception:
            p = ''
        if not p:
            return ''
        try:
            return hashlib.md5(p.encode('utf-8', errors='ignore')).hexdigest()
        except Exception:
            return ''

    def _qs_prj_key(self, key: str) -> str:
        pid = self._proj_id()
        if not pid:
            return ''
        return f"{self.SETTINGS_NS}/projects/{pid}/{key}"

    def _qs_prj_set(self, key: str, value) -> None:
        k = self._qs_prj_key(key)
        if not k:
            return
        try:
            self._qsettings.setValue(k, value)
        except Exception:
            pass

    def _qs_prj_get(self, key: str, default=None):
        k = self._qs_prj_key(key)
        if not k:
            return default
        try:
            return self._qsettings.value(k, default)
        except Exception:
            return default


    def _prj_get(self, key: str, default: str = "") -> str:
        prj = QgsProject.instance()
        ns = self.PROJECT_NS
        try:
            v, ok = prj.readEntry(ns, key, default)
            if ok:
                return v
        except Exception:
            pass
        try:
            return prj.readEntry(ns, key, default)[0]
        except Exception:
            return default

    def _prj_set(self, key: str, value: str) -> None:
        prj = QgsProject.instance()
        ns = self.PROJECT_NS
        try:
            prj.writeEntry(ns, key, str(value))
        except Exception:
            try:
                prj.writeEntry(ns, key, value)
            except Exception:
                pass


    def _prj_remove(self, key: str) -> None:
        prj = QgsProject.instance()
        ns = self.PROJECT_NS
        try:
            if hasattr(prj, 'removeEntry'):
                prj.removeEntry(ns, key)
            else:
                prj.writeEntry(ns, key, '')
        except Exception:
            try:
                prj.writeEntry(ns, key, '')
            except Exception:
                pass
    def _save_project_settings(self) -> None:
        # Do not dirty unsaved/empty projects. Only persist when project has a filename and user enabled persistence.
        if not self._project_can_persist():
            return
        try:
            try:
                self.sheet_name = self._cmb_sheet.currentText()
            except Exception:
                pass
            try:
                self.layer_name = self._cmb_layer.currentText()
            except Exception:
                pass
            try:
                self.layer_id = str(self._cmb_layer.currentData() or '')
            except Exception:
                self.layer_id = ''
            try:
                lyr = QgsProject.instance().mapLayer(self.layer_id) if self.layer_id else None
                self.layer_source = str(lyr.source()) if (lyr is not None and hasattr(lyr, 'source')) else ''
            except Exception:
                self.layer_source = ''

            try:
                self.layer_join_field = self._cmb_layer_key.currentText()
            except Exception:
                pass
            try:
                self.csv_join_field = self._cmb_csv_key.currentText()
            except Exception:
                pass

            self._prj_set('persist_enabled', '1')
            self._prj_set('excel_path', self.excel_path or '')
            self._prj_set('sheet_name', self.sheet_name or '')
            self._prj_set('csv_folder', self.csv_folder or '')
            self._prj_set('layer_name', self.layer_name or '')
            self._prj_set('layer_id', self.layer_id or '')
            self._prj_set('layer_source', self.layer_source or '')
            self._prj_set('layer_join_field', self.layer_join_field or '')
            self._prj_set('csv_join_field', self.csv_join_field or '')
            self._prj_set('interval_sec', str(self._get_interval_sec()))
            # fallback persistence to user settings (project-path keyed)
            try:
                self._qs_prj_set('excel_path', self.excel_path or '')
                self._qs_prj_set('sheet_name', self.sheet_name or '')
                self._qs_prj_set('csv_folder', self.csv_folder or '')
                self._qs_prj_set('layer_name', self.layer_name or '')
                self._qs_prj_set('layer_id', self.layer_id or '')
                self._qs_prj_set('layer_source', self.layer_source or '')
                self._qs_prj_set('layer_join_field', self.layer_join_field or '')
                self._qs_prj_set('csv_join_field', self.csv_join_field or '')
                self._qs_prj_set('interval_sec', str(self._get_interval_sec()))
            except Exception:
                pass

        except Exception:
            pass
    def _load_project_settings(self) -> None:
        self.excel_path = self._prj_get('excel_path', '') or ''
        self.sheet_name = self._prj_get('sheet_name', '') or ''
        self.csv_folder = self._prj_get('csv_folder', '') or ''
        self.layer_name = self._prj_get('layer_name', '') or ''
        self.layer_id = self._prj_get('layer_id', '') or ''
        self.layer_source = self._prj_get('layer_source', '') or ''
        self.layer_join_field = self._prj_get('layer_join_field', '') or ''
        self.csv_join_field = self._prj_get('csv_join_field', '') or ''
        try:
            interval = int(self._prj_get('interval_sec', '30') or '30')
        except Exception:
            interval = 30
        interval = 30 if interval < 45 else 60
        # fallback from user settings (project-path keyed) if project entries are missing
        try:
            if not (self.excel_path or self.csv_folder):
                self.excel_path = str(self._qs_prj_get('excel_path', self.excel_path or '') or '')
                self.sheet_name = str(self._qs_prj_get('sheet_name', self.sheet_name or '') or '')
                self.csv_folder = str(self._qs_prj_get('csv_folder', self.csv_folder or '') or '')
                self.layer_name = str(self._qs_prj_get('layer_name', self.layer_name or '') or '')
                self.layer_id = str(self._qs_prj_get('layer_id', self.layer_id or '') or '')
                self.layer_source = str(self._qs_prj_get('layer_source', self.layer_source or '') or '')
                self.layer_join_field = str(self._qs_prj_get('layer_join_field', self.layer_join_field or '') or '')
                self.csv_join_field = str(self._qs_prj_get('csv_join_field', self.csv_join_field or '') or '')
                try:
                    interval = int(str(self._qs_prj_get('interval_sec', interval) or interval))
                except Exception:
                    pass
        except Exception:
            pass
        try:
            self._spin_interval.setValue(interval)
        except Exception:
            pass
        try:
            self._txt_excel.setText(self.excel_path)
        except Exception:
            pass
        try:
            self._txt_csv.setText(self.csv_folder)
        except Exception:
            pass
        try:
            self.auto_show_dock = bool(self._qsettings.value(f"{self.SETTINGS_NS}/auto_show_dock", False, type=bool))
        except Exception:
            pass
    def _apply_state_to_ui(self):
        try:
            if hasattr(self, '_chk_persist'):
                # project opted-in shows checked, else reflect persist_enabled flag
                v = None
                try:
                    v = self._prj_get('persist_enabled', '0')
                except Exception:
                    v = None
                use_project = str(v) in ('1', 'true', 'True')
                self._chk_persist.setChecked(True if use_project else False)
        except Exception:
            pass

        if hasattr(self, "_txt_excel") and self._txt_excel:
            self._txt_excel.setText(self.excel_path or "")
        if hasattr(self, "_txt_csv") and self._txt_csv:
            self._txt_csv.setText(self.csv_folder or "")

        if self._cmb_sheet:
            try:
                self._populate_sheets()
            except Exception:
                pass
            if self.sheet_name:
                i = self._cmb_sheet.findText(self.sheet_name)
                if i >= 0:
                    self._cmb_sheet.setCurrentIndex(i)

        if self._cmb_layer:
            try:
                self._populate_layers()
            except Exception:
                pass
            if self.target_layer_id:
                for i in range(self._cmb_layer.count()):
                    if self._cmb_layer.itemData(i) == self.target_layer_id:
                        self._cmb_layer.setCurrentIndex(i)
                        break

        try:
            self._populate_layer_keys()
            self._populate_csv_keys()
        except Exception:
            pass

        if self._cmb_layer_key and self.layer_join_field:
            i = self._cmb_layer_key.findText(self.layer_join_field)
            if i >= 0:
                self._cmb_layer_key.setCurrentIndex(i)

        if self._cmb_csv_key and self.csv_join_field:
            i = self._cmb_csv_key.findText(self.csv_join_field)
            if i >= 0:
                self._cmb_csv_key.setCurrentIndex(i)

        if self._spin_interval:
            self._spin_interval.setValue(getattr(self, "_loaded_interval", 2))
            self._timer.setInterval(int(self._spin_interval.value()) * 1000)

        if self._chk_auto:
            self._chk_auto.setChecked(bool(getattr(self, "_loaded_auto", True)))
            if self._chk_auto.isChecked() and not self._timer.isActive():
                self._timer.start()
            if (not self._chk_auto.isChecked()) and self._timer.isActive():
                self._timer.stop()

    # ---------------- selections ----------------
    def select_excel(self):
        path, _ = QFileDialog.getOpenFileName(self.iface.mainWindow(), "Excelファイルを選択", "", "Excel (*.xlsx)")
        if not path:
            return
        self.excel_path = path
        self._txt_excel.setText(path)
        self._last_excel_mtime = self._safe_mtime(path)
        self._last_excel_size = self._safe_size(path)
        self._log(f"ui: select_excel -> {path}")

        self._populate_sheets()
        # first export to create minimal CSV for selecting join key
        self._schedule_export("select_excel")
        try:
            self._save_global_settings()
        except Exception:
            pass

    def select_csv_folder(self):
        folder = QFileDialog.getExistingDirectory(self.iface.mainWindow(), "CSV出力フォルダを選択", "")
        if not folder:
            return
        self.csv_folder = folder
        self._txt_csv.setText(folder)
        self.latest_csv = os.path.join(folder, "latest.csv")
        self.fallback_csv = os.path.join(folder, "fallback.csv")
        self._touch_watchlog()
        self._log(f"ui: select_out_folder -> {folder}")

        # If excel already selected, try export now to create latest.csv
        if self.excel_path:
            self._schedule_export("select_csv_folder")
        else:
            self._update_status("CSVフォルダ選択（Excel未選択）")
        try:
            self._save_global_settings()
        except Exception:
            pass

    def _on_sheet_changed(self, name: str):
        self.sheet_name = name.strip()
        if self.excel_path and self.csv_folder:
            # sheet changed => regenerate to reflect columns
            self._schedule_export("sheet_changed")

    # ---------------- populate combos ----------------
        try:
            self._save_global_settings()
        except Exception:
            pass
    def _populate_sheets(self):
        self._cmb_sheet.blockSignals(True)
        self._cmb_sheet.clear()
        names = []
        if self.excel_path and _HAVE_OPENPYXL:
            try:
                wb = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
                names = list(wb.sheetnames)
                wb.close()
            except Exception:
                names = []
        if not names and self.excel_path:
            # last resort: just show Sheet1
            names = ["Sheet1"]
        self._cmb_sheet.addItems(names)
        if self.sheet_name in names:
            self._cmb_sheet.setCurrentText(self.sheet_name)
        else:
            self.sheet_name = names[0] if names else ""
        self._cmb_sheet.blockSignals(False)

        # update csv keys based on current latest csv (if exists)
        self._populate_csv_keys()

    def _populate_layers(self):
        self._cmb_layer.blockSignals(True)
        self._cmb_layer.clear()
        items = []
        for lyr in QgsProject.instance().mapLayers().values():
            if isinstance(lyr, QgsVectorLayer):
                items.append((lyr.name(), lyr.id()))
        items.sort(key=lambda x: x[0].lower())
        for name, lid in items:
            self._cmb_layer.addItem(name, lid)
        self._cmb_layer.blockSignals(False)
        self._on_layer_changed()

    def _on_layer_changed(self):
        lid = self._cmb_layer.currentData()
        self.target_layer_id = lid or ""
        self._populate_layer_keys()
        try:
            self._save_global_settings()
        except Exception:
            pass

    def _populate_layer_keys(self):
        self._cmb_layer_key.clear()
        lyr = self._get_target_layer()
        if not lyr:
            return
        self._cmb_layer_key.addItems([f.name() for f in lyr.fields()])

        # default join key to OBJECTID when available
        try:
            if not self._cmb_layer_key.currentText():
                if self._cmb_layer_key.findText('OBJECTID') >= 0:
                    self._cmb_layer_key.setCurrentText('OBJECTID')
        except Exception:
            pass

        try:
            self._pick_default_join_field(self._cmb_layer_key)
        except Exception:
            pass

    def _populate_csv_keys(self):
        self._cmb_csv_key.clear()
        # Load latest csv layer (if exists) and list its fields
        lyr = self._load_csv_layer(silent=True)
        if lyr:
            self._cmb_csv_key.addItems([f.name() for f in lyr.fields()])


        # default CSV join key to OBJECTID when available
        try:
            if not self._cmb_csv_key.currentText():
                if self._cmb_csv_key.findText('OBJECTID') >= 0:
                    self._cmb_csv_key.setCurrentText('OBJECTID')
        except Exception:
            pass

        try:
            self._pick_default_join_field(self._cmb_csv_key)
        except Exception:
            pass
    # ---------------- auto settings ----------------
    def _on_auto_changed(self):
        if self._chk_auto.isChecked():
            if not self._timer.isActive():
                self._timer.start()
            self._update_status("自動同期ON")
        else:
            self._timer.stop()
            self._update_status("自動同期OFF")

    def _on_join_keys_changed(self, *_args):
        try:
            self.layer_join_field = self._cmb_layer_key.currentText().strip() if self._cmb_layer_key else ""
            self.csv_join_field = self._cmb_csv_key.currentText().strip() if self._cmb_csv_key else ""
        except Exception:
            pass
        try:
            self._save_project_settings()
        except Exception:
            pass
        try:
            self._save_global_settings()
        except Exception:
            pass


    def _on_interval_changed(self, v: int):
        self._timer.setInterval(int(v) * 1000)

    # ---------------- logging / status ----------------
        try:
            self._save_global_settings()
        except Exception:
            pass
    def _touch_watchlog(self):
        # make sure log exists in csv folder
        if not self.csv_folder:
            return
        path = os.path.join(self.csv_folder, "watch_log.txt")
        try:
            with open(path, "a", encoding="utf-8") as f:
                f.write(f"{self._ts()}\tTOUCH\n")
        except Exception:
            pass

    def _log(self, msg: str):
        if not self.csv_folder:
            return
        path = os.path.join(self.csv_folder, "watch_log.txt")
        try:
            with open(path, "a", encoding="utf-8") as f:
                f.write(f"{self._ts()}\t{msg}\n")
        except Exception:
            # if even that fails, we can't do more safely
            pass

    def _update_status(self, note: str):
        excel_m = self._safe_mtime(self.excel_path) if self.excel_path else 0.0
        csv_m = self._safe_mtime(self.latest_csv) if self.latest_csv else 0.0
        csv_s = self._safe_size(self.latest_csv) if self.latest_csv else -1
        msg = (
            f"状態: {note} / "
            f"Excel更新: {self._fmt_ts(excel_m) if excel_m else '-'} / "
            f"CSV更新: {self._fmt_ts(csv_m) if csv_m else '-'} / "
            f"CSVサイズ: {csv_s if csv_s >= 0 else 'None'}"
        )
        if self._lbl_status:
            self._lbl_status.setText(msg)

    # ---------------- polling / trigger ----------------
    def _poll_updates(self):
        # called by QTimer
        if not (self._chk_auto and self._chk_auto.isChecked()):
            return

        # 1) Excel changed?
        if self.excel_path and self.csv_folder:
            m = self._safe_mtime(self.excel_path)
            s = self._safe_size(self.excel_path)
            if (m and m != self._last_excel_mtime) or (s >= 0 and s != self._last_excel_size):
                self._log("poll: excel changed -> schedule export")
                self._last_excel_mtime = m
                self._last_excel_size = s
                self._schedule_export("poll_excel_changed")

        # 2) CSV changed? (e.g. external process)
        if self.latest_csv:
            cm = self._safe_mtime(self.latest_csv)
            cs = self._safe_size(self.latest_csv)
            if (cm and cm != self._last_csv_mtime) or (cs >= 0 and cs != self._last_csv_size):
                self._last_csv_mtime = cm
                self._last_csv_size = cs
                self._log("poll: csv changed -> refresh join fields / apply join")
                self._populate_csv_keys()
                # try join
                self.execute_virtual_join(quiet=True)

        self._update_status("監視中")


    def _wait_file_readable(self, path: str, timeout_s: float = 8.0) -> bool:
        """Wait until file can be opened for reading (Windows Excel lock guard)."""
        if not path:
            return False
        t0 = time.time()
        while time.time() - t0 < timeout_s:
            try:
                with open(path, "rb"):
                    return True
            except PermissionError:
                time.sleep(0.3)
            except OSError:
                time.sleep(0.3)
        return False

    def _wait_file_stable(self, path: str, checks: int = 2, interval_ms: int = 250) -> bool:
        """Return True if file exists, is non-empty, and size is stable across checks."""
        try:
            if not path or (not os.path.exists(path)):
                return False
            last = os.path.getsize(path)
            if last <= 0:
                return False
            for _ in range(checks):
                time.sleep(interval_ms / 1000.0)
                cur = os.path.getsize(path)
                if cur != last or cur <= 0:
                    return False
                last = cur
            return True
        except Exception:
            return False


    def _get_interval_sec(self) -> int:
        """Return watch interval snapped to 30 or 60 seconds."""
        try:
            v = self._get_interval_sec()
        except Exception:
            v = 30
        return 30 if v < 45 else 60

    def _schedule_export(self, reason: str):
        # reset retry counter for a new export request
        self._export_retry_count = 0
        if not (self.excel_path and self.csv_folder and self.sheet_name):
            self._update_status("Excel/CSV/シート未設定")
            return
        self._log(f"schedule_export: {reason}")
        if self._pending_export:
            return
        self._pending_export = True
        QTimer.singleShot(800, self._do_export)

    # ---------------- export ----------------
    def _do_export(self):
        self._pending_export = False
        ok = False
        err = ""
        try:
            ok = self.export_excel_to_csv()
        except RuntimeError as e:
            if str(e) == '__LOCKED__':
                # Excel is still locked (often during Save / OneDrive sync). Retry with backoff.
                self._export_retry_count += 1
                if self._export_retry_count <= self._export_retry_max:
                    delay_ms = min(30000, 800 + self._export_retry_count * 1200)  # 2s..max 30s
                    self._log(f"export: Excel locked -> retry {self._export_retry_count}/{self._export_retry_max} in {int(delay_ms/1000)}s")
                    QTimer.singleShot(delay_ms, self._do_export)
                    return
                err = "Excel file locked"
                ok = False
            else:
                raise
        except Exception as e:
            err = str(e)
            ok = False

        if ok:
            self._log("export: OK")
            # Update last signatures
            try:
                self._last_excel_mtime = self._safe_mtime(self.excel_path)
                self._last_excel_size = self._safe_size(self.excel_path)
            except Exception:
                pass
            try:
                self._last_csv_hash = self._csv_file_hash(self.latest_csv)
            except Exception:
                pass

            # Diagnostics: verify that trailing text columns (e.g. R8) are really present in the exported CSV.
            try:
                import csv as _csv
                with open(self.csv_path, "r", encoding="utf-8-sig", newline="") as _f:
                    _reader = _csv.DictReader(_f)
                    _fields = _reader.fieldnames or []
                    if "R8" in _fields:
                        _t = 0
                        _n = 0
                        for _row in _reader:
                            _t += 1
                            if str(_row.get("R8", "")).strip():
                                _n += 1
                            if _t >= 5000:
                                break
                        self._log(f"diag: CSV R8 nonempty {_n}/{_t}")
            except Exception:
                pass
            self._populate_csv_keys()
            self._update_status("CSV生成OK")
            # apply join after export (deferred)
            def _do_join():
                try:
                    if bool(getattr(self, '_is_restoring', False)):
                        QTimer.singleShot(150, _do_join)
                        return
                    self.execute_virtual_join(quiet=True)
                except Exception:
                    pass
            QTimer.singleShot(0, _do_join)
        else:
            self._log("export: FAILED " + (err or ""))
            self._update_status("CSV生成失敗")
    def export_excel_to_csv(self) -> bool:
        """Write latest.csv and fallback.csv. Return True if latest.csv is created and non-empty."""
        if not (self.excel_path and self.csv_folder and self.sheet_name):
            return False

        os.makedirs(self.csv_folder, exist_ok=True)
        latest = self.latest_csv or os.path.join(self.csv_folder, "latest.csv")
        fallback = self.fallback_csv or os.path.join(self.csv_folder, "fallback.csv")
        # Guard: wait until Excel is readable and stable (prevents reading during save/lock)
        self._wait_file_readable(self.excel_path, timeout_s=8.0)
        self._wait_file_stable(self.excel_path, checks=2, interval_ms=250)


        # 1) primary: openpyxl (keeps Excel text "as-is"; avoids phonetic-kana)
        if _HAVE_OPENPYXL:
            try:
                wb = None
                for _i in range(6):
                    try:
                        wb = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
                        break
                    except PermissionError:
                        time.sleep(0.4)
                    except OSError:
                        time.sleep(0.4)
                if wb is None:
                    raise RuntimeError('__LOCKED__')
                if self.sheet_name in wb.sheetnames:
                    ws = wb[self.sheet_name]
                else:
                    ws = wb[wb.sheetnames[0]]
                    self.sheet_name = ws.title
                with open(latest, "w", encoding="utf-8-sig", newline="") as f:
                    w = csv.writer(f)

                    # Determine column range from the first few rows (Excel used-range can be smaller than expected).
                    # This avoids silently dropping trailing columns like "R8".
                    max_col = 0
                    for rr in range(1, min(5, ws.max_row) + 1):
                        row_cells = ws[rr]
                        for cc, cell in enumerate(row_cells, start=1):
                            v = cell.value
                            if v is None:
                                continue
                            if isinstance(v, str) and v.strip() == "":
                                continue
                            max_col = max(max_col, cc)
                    if max_col <= 0:
                        max_col = ws.max_column

                    # Export rows with cell objects so we can fall back when data_only yields None.
                    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col, values_only=False):
                        out = []
                        for cell in row_cells:
                            v = cell.value
                            if v is None:
                                # If it is a formula and cache isn't present, keep empty instead of "None".
                                out.append("")
                            else:
                                out.append(str(v))
                        w.writerow(out)
                wb.close()
                # if file exists and has size
                if self._safe_size(latest) > 0:
                    self._copy_file(latest, fallback)
                    self._last_csv_mtime = self._safe_mtime(latest)
                    self._last_csv_size = self._safe_size(latest)
                    return True
            except Exception as e:
                self._log("export(openpyxl) failed: " + str(e))

        # 2) fallback: OGR reading through QGIS + write to CSV
        try:
            vl = self._load_excel_as_layer()
            if not vl or not vl.isValid():
                self._log("export(ogr) failed: cannot open excel as vector layer")
                return False

            opts = QgsVectorFileWriter.SaveVectorOptions()
            opts.driverName = "CSV"
            opts.fileEncoding = "UTF-8"
            ctx = QgsCoordinateTransformContext()
            res, _, _ = QgsVectorFileWriter.writeAsVectorFormatV2(vl, latest, ctx, opts)
            if res != QgsVectorFileWriter.NoError:
                self._log(f"export(ogr) failed: writer error code {res}")
                return False

            if self._safe_size(latest) > 0:
                self._copy_file(latest, fallback)
                self._last_csv_mtime = self._safe_mtime(latest)
                self._last_csv_size = self._safe_size(latest)
                return True
        except Exception as e:
            self._log("export(ogr) exception: " + str(e))
            self._log(traceback.format_exc())


        # If export failed but an existing latest.csv is present and fresh, accept it to avoid false 'failed' status.
        try:
            if os.path.exists(latest) and self._safe_size(latest) > 0:
                excel_m = self._safe_mtime(self.excel_path)
                csv_m = self._safe_mtime(latest)
                if csv_m >= (excel_m - 5):
                    self._log('export: using existing latest.csv (fresh enough)')
                    return True
        except Exception:
            pass

        return False
    def _load_excel_as_layer(self) -> Optional[QgsVectorLayer]:
        """Try to open Excel sheet as vector layer via OGR provider."""
        # Typical OGR uri: "path.xlsx|layername=Sheet1"
        if not self.excel_path:
            return None
        sheet = self.sheet_name or "Sheet1"
        uri = f"{self.excel_path}|layername={sheet}"
        vl = QgsVectorLayer(uri, "excel_sheet", "ogr")
        if vl.isValid():
            # Debug: list CSV fields
            try:
                self._log("excel sheet fields: " + ", ".join([f.name() for f in vl.fields()]))
            except Exception:
                pass
            return vl
        # try without layername (first sheet)
        vl = QgsVectorLayer(self.excel_path, "excel_sheet", "ogr")
        return vl if vl.isValid() else None

    def _load_csv_layer(self, silent: bool = False) -> Optional[QgsVectorLayer]:
        """Load latest.csv as a delimited text layer and keep it in the project.

        Important:
          - If the join layer is not in QgsProject, QGIS may not expose joined fields in attribute table.
          - We also keep a Python reference to avoid GC while joins are active.

        Note:
          - The delimitedtext provider may keep old contents if we reuse the same QgsVectorLayer object
            while the CSV file is overwritten. We therefore recreate the layer when the CSV mtime changes.
        """
        if not (self.latest_csv and os.path.exists(self.latest_csv)):
            return None

        csv_mtime = self._safe_mtime(self.latest_csv)
        csv_size = self._safe_size(self.latest_csv)

        # Reuse cached layer only if it matches the file AND the mtime did not change
        try:
            if getattr(self, "_csv_layer", None) is not None:
                same_file = self._csv_layer.isValid() and (self.latest_csv in self._csv_layer.source())
                same_mtime = (getattr(self, "_csv_layer_mtime", None) == csv_mtime)
                same_size = (getattr(self, "_csv_layer_size", None) == csv_size)
                if same_file and same_mtime and same_size:
                    return self._csv_layer

                # Remove old layer from project to avoid duplicate "csv_latest"
                try:
                    QgsProject.instance().removeMapLayer(self._csv_layer.id())
                except Exception:
                    pass

                self._csv_layer = None
        except Exception:
            self._csv_layer = None

        uri = self._make_delimitedtext_uri(self.latest_csv)
        vl = QgsVectorLayer(uri, "csv_latest", "delimitedtext")
        if not vl.isValid():
            if not silent:
                self._log("csv layer invalid: " + uri)
            return None

        # Add to project without showing in layer tree (False)
        try:
            QgsProject.instance().addMapLayer(vl, False)
        except Exception:
            # Even if add fails, keep the layer; join may still work in some cases
            pass

        self._csv_layer = vl
        self._csv_layer_mtime = csv_mtime
        self._csv_layer_size = csv_size
        return vl

    def _make_delimitedtext_uri(self, path: str) -> str:
        # file:///C:/path/latest.csv?delimiter=,&quote="&escape="&header=Yes&detectTypes=no&charset=UTF-8
        p = path.replace("\\", "/")
        if not p.lower().startswith("/"):
            # windows drive
            p = "/" + p
        uri = f"file://{p}?delimiter=,&quote=\"&escape=\"&header=Yes&detectTypes=no&charset=UTF-8"
        return uri

    # ---------------- join ----------------

    def _pick_default_join_field(self, cmb) -> None:
        """Prefer OBJECTID; otherwise first non-fid field."""
        try:
            cur = (cmb.currentText() or '').strip()
        except Exception:
            cur = ''
        if cur.lower() == 'fid':
            cur = ''
        try:
            if not cur and cmb.findText('OBJECTID') >= 0:
                cmb.setCurrentText('OBJECTID')
                return
        except Exception:
            pass
        try:
            if not cur:
                for i in range(cmb.count()):
                    t = (cmb.itemText(i) or '').strip()
                    if t and t.lower() != 'fid':
                        cmb.setCurrentText(t)
                        return
        except Exception:
            pass

    def _set_chk_checked_blocked(self, chk, checked: bool) -> None:
        try:
            if chk is None:
                return
            chk.blockSignals(True)
            chk.setChecked(bool(checked))
        except Exception:
            pass
        finally:
            try:
                chk.blockSignals(False)
            except Exception:
                pass

    def _ensure_string_join_field(self, layer, field_name: str) -> str:
        """Return original field name. Also remove any legacy __join_str_* helper expression fields if present."""
        try:
            self._remove_join_helper_fields(layer)
        except Exception:
            pass
        return field_name

    def _remove_join_helper_fields(self, layer):
        """Remove legacy expression fields like __join_str_* created by older versions."""
        if not layer or (not layer.isValid()):
            return
        try:
            # Expression fields are appended; remove by name match
            helper_re = re.compile(r"^__join_str_.*$", re.IGNORECASE)
            fields = layer.fields()
            # collect indices first
            to_remove = []
            for i, f in enumerate(fields):
                if helper_re.match(f.name()):
                    to_remove.append(i)
            # remove in reverse order
            for i in reversed(to_remove):
                try:
                    layer.removeExpressionField(i)
                except Exception:
                    pass
        except Exception:
            pass
    def _hide_helper_fields(self, layer: QgsVectorLayer):
        """Hide internal helper fields such as __join_str_* from attribute table and forms."""
        if not layer or (not layer.isValid()):
            return
        helper_re = re.compile(r"^_+join_str_.*$", re.IGNORECASE)

        try:
            cfg = layer.attributeTableConfig()
        except Exception:
            cfg = None

        for i, f in enumerate(layer.fields()):
            name = f.name()
            if not helper_re.match(name):
                continue

            # 1) Hide in attribute table
            try:
                if cfg is not None:
                    cfg.setColumnHidden(i, True)
            except Exception:
                pass

            # 2) Hide in edit form
            try:
                layer.setEditorWidgetSetup(i, QgsEditorWidgetSetup("Hidden", {}))
            except Exception:
                pass

        try:
            if cfg is not None:
                layer.setAttributeTableConfig(cfg)
        except Exception:
            pass

        try:
            layer.triggerRepaint()
        except Exception:
            pass



    def _csv_file_hash(self, csv_path: str) -> str:
        """Hash CSV content to avoid false-change loops (mtime changes even if content same)."""
        try:
            h = hashlib.md5()
            with open(csv_path, 'rb') as f:
                for chunk in iter(lambda: f.read(1024 * 1024), b''):
                    h.update(chunk)
            return h.hexdigest()
        except Exception:
            return ''

    def _csv_schema_hash(self, csv_path: str) -> str:
        try:
            with open(csv_path, 'r', encoding='utf-8-sig', newline='') as f:
                reader = csv.reader(f)
                header = next(reader, None) or []
            h = '|'.join([str(x).strip() for x in header])
            return hashlib.md5(h.encode('utf-8', errors='ignore')).hexdigest()
        except Exception:
            return ''

    def _apply_csv_field_order(self, layer: QgsVectorLayer) -> None:
        """Reorder attribute table columns so joined csv_* fields follow CSV header order."""
        try:
            csv_path = getattr(self, "csv_path", "") or getattr(self, "latest_csv", "") or ""
            if (not csv_path) or (not os.path.exists(csv_path)):
                return
            with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.reader(f)
                header = next(reader, None) or []
            header = [h.strip() for h in header if h is not None]
            if not header:
                return
            desired_join = [f"csv_{h}" for h in header if h]
            all_fields = [fld.name() for fld in layer.fields()]
            non_join = [n for n in all_fields if not n.startswith("csv_")]
            join_present = [n for n in desired_join if n in all_fields]
            remaining = [n for n in all_fields if (n.startswith("csv_") and n not in join_present)]
            final_names = non_join + join_present + remaining

            cfg = layer.attributeTableConfig()
            cols_by_name = {c.name: c for c in cfg.columns()}
            new_cols = []
            for name in final_names:
                if name in cols_by_name:
                    new_cols.append(cols_by_name[name])
                else:
                    new_cols.append(QgsAttributeTableConfig.ColumnConfig(name))
            cfg.setColumns(new_cols)
            layer.setAttributeTableConfig(cfg)
        except Exception:
            pass

    def execute_virtual_join(self, quiet: bool = True) -> bool:
        if bool(getattr(self, '_is_restoring', False)):
            return False
        # Speed: skip join when latest.csv is unchanged
        try:
            csv_path = (self.latest_csv or '').strip()
            mtime = self._safe_mtime(csv_path) if csv_path else 0.0
            size = self._safe_size(csv_path) if csv_path else 0
            schema = self._csv_schema_hash(csv_path) if csv_path else ''
            if mtime == self._last_csv_mtime and size == self._last_csv_size and (not schema or schema == self._last_schema_hash):
                self._log('join: skip (csv unchanged)')
                return True
            self._last_csv_mtime = mtime
            self._last_csv_size = size
            if schema:
                self._last_schema_hash = schema
        except Exception:
            pass
        lyr = self._get_target_layer()
        if not lyr:
            if not quiet:
                QMessageBox.warning(self.iface.mainWindow(), "Join", "対象レイヤが選択されていません。")
            return False
        if not self.latest_csv or not os.path.exists(self.latest_csv):
            if not quiet:
                QMessageBox.warning(self.iface.mainWindow(), "Join", "latest.csv がありません。先にCSV生成してください。")
            return False

        # avoid accidental fid join: pick good defaults
        try:
            self._pick_default_join_field(self._cmb_layer_key)
        except Exception:
            pass
        try:
            self._pick_default_join_field(self._cmb_csv_key)
        except Exception:
            pass

        layer_key = self._cmb_layer_key.currentText().strip() if self._cmb_layer_key else ""
        csv_key = self._cmb_csv_key.currentText().strip() if self._cmb_csv_key else ""
        # If CSV expects OBJECTID but selected layer has no OBJECTID key, skip (prevents wrong-layer 0% joins).
        try:
            if (csv_key or '').strip() == 'OBJECTID' and (layer_key or '').strip() != 'OBJECTID':
                self._log(f"join: skipped (layer has no OBJECTID) layer_key={layer_key} csv_key={csv_key}")
                return False
        except Exception:
            pass

        self.layer_join_field = layer_key
        self.csv_join_field = csv_key
        try:
            self._save_project_settings()
        except Exception:
            pass
        if not (layer_key and csv_key):
            if not quiet:
                QMessageBox.warning(self.iface.mainWindow(), "Join", "Joinキーを選択してください。")
            return False

        csv_layer = self._load_csv_layer(silent=True)
        if not csv_layer:
            if not quiet:
                QMessageBox.warning(self.iface.mainWindow(), "Join", "CSVレイヤを読み込めません。")
            return False

        # remove old joins
        self._remove_joins_by_prefix(lyr, prefix="csv_latest")
        self._remove_joins_by_prefix(lyr, prefix="excelcsv_")

        # add join
        join = QgsVectorLayerJoinInfo()
        join.setJoinLayer(csv_layer)
        join.setJoinFieldName(csv_key)
        layer_key2 = self._ensure_string_join_field(lyr, layer_key)
        join.setTargetFieldName(layer_key2)
        join.setUsingMemoryCache(False)
        join.setPrefix("csv_")
        # JOIN all CSV columns: do not limit subset (QGIS will join all fields by default)

        lyr.addJoin(join)
        # align attribute table columns to CSV order
        self._apply_csv_field_order(lyr)

        # hide helper fields again (join refresh may reset visibility)
        try:
            self._hide_helper_fields(lyr)
        except Exception:
            pass

        self._log(f"join: applied {lyr.name()} ({layer_key2}) <- CSV ({csv_key})")

        try:

            matched, total, ex = self._calc_join_success_rate(lyr, layer_key2, csv_layer, csv_key, sample=5)

            if total > 0:

                pct = (matched / total) * 100.0

                self._log(f"join: match {matched}/{total} ({pct:.1f}%)")

                if ex:

                    self._log("join: example unmatched keys (target): " + ", ".join(ex))

        except Exception as e:

            self._log(f"join: match calc error: {e}")
        self._update_status("Join更新")
        return True

    def remove_existing_joins(self):
        lyr = self._get_target_layer()
        if not lyr:
            return
        self._remove_joins_by_prefix(lyr, prefix="csv_")
        self._remove_joins_by_prefix(lyr, prefix="excelcsv_")
        self._update_status("Join解除")

    def _remove_joins_by_prefix(self, layer: QgsVectorLayer, prefix: str):
        try:
            joins = list(layer.vectorJoins())
            for j in joins:
                try:
                    if getattr(j, "prefix", "") and j.prefix().startswith(prefix):
                        layer.removeJoin(j.joinLayerId())
                except Exception:
                    # remove anyway if join layer name matches
                    layer.removeJoin(j.joinLayerId())
        except Exception:
            pass

    def _norm_join_key(self, v):
        """Normalize join key to improve match robustness (string/number/None)."""
        if v is None:
            return ""
        try:
            # QVariant -> python value
            if hasattr(v, "toPyObject"):
                v = v.toPyObject()
        except Exception:
            pass
        try:
            s = str(v)
        except Exception:
            return ""
        s = s.strip()
        # common numeric artifacts like '123.0'
        if re.match(r"^-?\d+\.0$", s):
            s = s[:-2]
        return s

    def _calc_join_success_rate(self, target_layer, target_key_field, csv_layer, csv_key_field, sample=5):
        """Return (matched, total, examples) where examples shows a few unmatched keys."""
        matched = 0
        total = 0
        examples = []
        try:
            csv_keys = set()
            if csv_layer is None:
                return 0, 0, []
            for f in csv_layer.getFeatures():
                try:
                    csv_keys.add(self._norm_join_key(f[csv_key_field]))
                except Exception:
                    continue

            for f in target_layer.getFeatures():
                total += 1
                try:
                    k = self._norm_join_key(f[target_key_field])
                except Exception:
                    k = ""
                if k in csv_keys:
                    matched += 1
                else:
                    if len(examples) < sample and k != "":
                        examples.append(k)
        except Exception as e:
            self._log(f"join: success_rate calc failed: {e}")
        return matched, total, examples

    def _get_target_layer(self) -> Optional[QgsVectorLayer]:
        if not self.target_layer_id:
            return None
        lyr = QgsProject.instance().mapLayer(self.target_layer_id)
        return lyr if isinstance(lyr, QgsVectorLayer) else None


    def _apply_csv_folder_runtime(self) -> None:
        """Ensure internal runtime paths for csv_folder are initialized."""
        try:
            folder = getattr(self, 'csv_folder', '') or ''
            if not folder:
                return
            self.latest_csv = os.path.join(folder, 'latest.csv')
            self.fallback_csv = os.path.join(folder, 'fallback.csv')
        except Exception:
            pass
        try:
            if getattr(self, 'csv_folder', ''):
                self._touch_watchlog()
        except Exception:
            pass

    def _restore_layer_selection_c(self) -> bool:
        """Restore layer selection by (id -> source -> name)."""
        try:
            if self._cmb_layer is None or self._cmb_layer.count() == 0:
                return False
        except Exception:
            return False

        wanted_id = getattr(self, 'layer_id', '') or ''
        wanted_source = getattr(self, 'layer_source', '') or ''
        wanted_name = getattr(self, 'layer_name', '') or ''

        if wanted_id:
            try:
                for i in range(self._cmb_layer.count()):
                    if str(self._cmb_layer.itemData(i) or '') == wanted_id:
                        self._cmb_layer.setCurrentIndex(i)
                        return True
            except Exception:
                pass

        if wanted_source:
            try:
                prj = QgsProject.instance()
                for i in range(self._cmb_layer.count()):
                    lid = str(self._cmb_layer.itemData(i) or '')
                    lyr = prj.mapLayer(lid) if lid else None
                    if lyr is not None and hasattr(lyr, 'source') and str(lyr.source()) == wanted_source:
                        self._cmb_layer.setCurrentIndex(i)
                        return True
            except Exception:
                pass

        if wanted_name:
            try:
                idx = self._cmb_layer.findText(wanted_name)
                if idx >= 0:
                    self._cmb_layer.setCurrentIndex(idx)
                    return True
            except Exception:
                pass

        return False

    def _restore_layer_selection_c_deferred(self, tries: int = 8) -> None:
        try:
            if self._restore_layer_selection_c():
                return
        except Exception:
            pass
        if tries <= 0:
            return
        try:
            QTimer.singleShot(250, lambda: self._restore_layer_selection_c_deferred(tries - 1))
        except Exception:
            pass


    # ---------------- helpers ----------------
    def _open_latest_csv(self):
        if not self.latest_csv or not os.path.exists(self.latest_csv):
            QMessageBox.information(self.iface.mainWindow(), "CSV", "latest.csv がありません。")
            return
        # open with default app (Windows)
        try:
            import subprocess
            subprocess.Popen(["cmd", "/c", "start", "", self.latest_csv], shell=True)
        except Exception:
            pass

    def _safe_mtime(self, path: str) -> float:
        try:
            return os.path.getmtime(path) if path and os.path.exists(path) else 0.0
        except Exception:
            return 0.0

    def _safe_size(self, path: str) -> int:
        try:
            return os.path.getsize(path) if path and os.path.exists(path) else -1
        except Exception:
            return -1

    def _copy_file(self, src: str, dst: str):
        try:
            import shutil
            shutil.copy2(src, dst)
        except Exception:
            pass

    def _ts(self) -> str:
        return time.strftime("%Y-%m-%d %H:%M:%S")

    def _fmt_ts(self, mtime: float) -> str:
        try:
            if not mtime:
                return "-"
            return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(mtime))
        except Exception:
            return "-"
